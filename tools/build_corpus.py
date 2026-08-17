"""语料生成器:解析桌面 Excel → corpus/*.jsonl + config/banned_keywords.json。

脱敏守卫(重要):
- FORBIDDEN_SHEETS 硬编码(定价/价卡/刷价,含真实 SKU/成本),绝不打开;
- 生成内容正则扫 6 位以上数字 / Seller-SKU 模式,命中即丢弃并记入 parse_report;
- 原始 Excel 不入仓库(.gitignore 已排除 *.xlsx)。

用法:
    python tools/build_corpus.py
    python tools/build_corpus.py --src "C:\\Users\\zwx\\Desktop\\新建文件夹"

输出:
    corpus/qa_faq.jsonl · qa_rules.jsonl · qa_pricing.jsonl
    config/banned_keywords.json
    data/parse_report.md   (人工复核存疑行)
"""
import argparse
import json
import re
import sys
from collections import OrderedDict, defaultdict
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

BASE = Path(__file__).resolve().parent.parent
CORPUS_DIR = BASE / "corpus"
CONFIG_DIR = BASE / "config"
DATA_DIR = BASE / "data"

DEFAULT_SRC = Path(r"C:\Users\zwx\Desktop\新建文件夹")
TALKBOOK = "常用话术参考模板参考.xlsx"   # 话术:快捷回复/常见问题/退货退款/违禁表
PRICING = "原表5.xlsx"                  # 定价税费:只聚合,不出明细
FORBIDDEN_SHEETS = {"定价", "价卡", "刷价"}  # 真实经营数据,永不打开

# 敏感模式:6 位以上数字 / Seller-SKU / 带序号的 SKU 串
SENSITIVE_RE = re.compile(r"\b\d{6,}\b|[A-Za-z]{2,}[-\s]?\d{6,}|SKU[:：]?\s*\S{4,}")
_TITLE_RE = re.compile(r"^\s*\d+[.、．]")


# ---------------------------------------------------------------- 工具
def _cells(row):
    """openpyxl 行 → 去空白字符串列表(容忍 None/非文本)。"""
    return ["" if c is None else str(c).strip() for c in row]


def _guard(text, report, tag):
    """命中敏感模式返回 None(丢弃并记报告),否则原样返回。"""
    if text is None:
        return text
    if SENSITIVE_RE.search(text):
        report.append(f"[DROP:{tag}] 命中敏感模式: {text[:60]}")
        return None
    return text


def _is_title(text):
    """粗略判断小标题:编号开头 / 短句(≤12 字,不以句号/逗号结尾)。"""
    if not text:
        return False
    if _TITLE_RE.match(text):
        return True
    if len(text) <= 12 and not re.search(r"[。？！，,；;]$", text) \
            and not re.match(r"^\d+$", text):
        return True
    return False


def _add(entries, seq, prefix, question, answer, category, tags, source,
         report, lang="zh", markets=None):
    q = _guard(question, report, "question")
    a = _guard(answer, report, "answer")
    if not q or not a:
        return
    entries.append({
        "id": "%s-%04d" % (prefix, seq[0]), "question": q, "answer": a,
        "category": category, "tags": tags, "source": source,
        "lang": lang, "markets": markets or [],
    })
    seq[0] += 1


# ---------------------------------------------------------------- 各 sheet 解析
def parse_quick_reply(rows, report):
    """快捷回复:分类|场景|内容(分类/场景向前填充)。每行内容一条。

    rows: 已内存化的行元组列表(含表头行,解析时跳过)。
    """
    entries, seq = [], [0]
    cat = scene = ""
    for row in rows[1:]:
        c = _cells(row)
        if len(c) < 3:
            continue
        if c[0]:
            cat = c[0]
        if c[1]:
            scene = c[1]
        content = _guard(c[2], report, "快捷回复")
        if not content:
            continue
        q = scene or cat or "快捷回复"
        tags = [t for t in (cat, scene, "快捷回复") if t]
        _add(entries, seq, "faq", q, content, "快捷回复", tags,
             "话术库/快捷回复", report)
    return entries


def parse_faq_assistant(rows, report):
    """常见问题助理:商品|内容|答案。答案跨行合并;内嵌分区头跳过。"""
    entries, seq = [], [0]
    bucket = ""
    cur_q = cur_a = ""
    for row in rows[1:]:
        c = _cells(row)
        if len(c) < 3:
            continue
        prod, q, a = c[0], c[1], c[2]
        if q in ("内容", "问题", "答案"):  # 重复表头/分区头
            if prod and prod not in ("商品", "订单", "内容"):
                bucket = prod
            continue
        if q:  # 新问题
            if cur_q and cur_a:
                _add(entries, seq, "faq", cur_q, cur_a, "常见问题",
                     [b for b in (bucket,) if b], "话术库/常见问题助理", report)
            cur_q, cur_a = q, a
        elif a:  # 答案续行
            cur_a = (cur_a + " " + a).strip()
    if cur_q and cur_a:
        _add(entries, seq, "faq", cur_q, cur_a, "常见问题",
             [b for b in (bucket,) if b], "话术库/常见问题助理", report)
    return entries


def parse_customer_questions(rows, report):
    """客服的问题汇总:单列 B 流。小标题 → 分区,标题下内容合并为答案。"""
    entries, seq = [], [0]
    topic, buf = "", []
    def flush():
        if topic and buf:
            text = " ".join(buf).strip()
            _add(entries, seq, "faq", topic, text, "常见问题",
                 [topic], "话术库/客服的问题汇总", report)
        buf.clear()
    for row in rows:
        c = _cells(row)
        text = c[1] if len(c) > 1 else ""
        if not text:
            continue
        if _is_title(text):
            flush()
            topic = text
        else:
            g = _guard(text, report, "客服问题汇总")
            if g:
                buf.append(g)
    flush()
    return entries


def parse_refund(rows, report):
    """退货退款:列 B 分区头(订单未发货/拒绝退款)+ 分区下编号内容。按分区合并。"""
    entries, seq = [], [0]
    section, buf = "", []
    def flush():
        if section and buf:
            text = " ".join(buf).strip()
            _add(entries, seq, "faq", section + "怎么办？", text, "退货退款",
                 [section], "话术库/退货退款", report)
        buf.clear()
    for row in rows:
        c = _cells(row)
        text = c[1] if len(c) > 1 else ""
        if not text:
            continue
        if not c[0] and _is_title(text):
            flush()
            section = text
        else:
            g = _guard(text, report, "退货退款")
            if g:
                buf.append(g)
    flush()
    return entries


# 战术列白名单 + 战术特征正则(列 H 错位严重,只收明显战术词,其余进审查)
TACTICAL_WHITELIST = {"防弹", "子弹", "折叠刀", "军刀", "匕首", "仿真枪",
                      "弩", "甩棍", "指虎", "剑", "斧", "弓箭"}
TACTICAL_RE = re.compile(r"刀|弹|枪|剑|斧|箭|弩|棍|锤|匕|刺|刃|雷|炸|爆|火药|军火|工兵")
# 高频关键词列(E)里太泛的词,收录会误杀正常商品(如"成分"命中的护肤标题)
STOP_WORDS = {"成分", "目录", "伦理", "毒性", "挥发", "传染"}


def parse_banned(rows, report):
    """违禁自动刊登关键词 → (规则问答, 违禁词表)。

    列 A-E(分类/子类/限制/示例/高频关键词)→ 规则问答 + keywords;
    列 H(战术)→ 明显战术词进 free_keywords(战术列错位混入正常商品词,
    如 指甲油/花露水,剔除并记入报告待人工复核),再并入基础战术词兜底。
    """
    entries, seq = [], [0]
    keywords, free, cats, review = [], [], [], []
    for row in rows[1:]:
        c = _cells(row)
        if len(c) < 5:
            continue
        cat, sub, limit, examples, hf = c[0], c[1], c[2], c[3], c[4]
        for tok in re.split(r"[\s,，、;；/]+", hf):
            if tok and len(tok) >= 2 and tok not in STOP_WORDS and tok not in keywords:
                keywords.append(tok)
        if cat and cat not in cats:
            cats.append({"cat": cat, "sub": sub, "limit": limit,
                         "examples": examples})
        if len(c) >= 8 and c[7]:
            t = c[7]
            if t in TACTICAL_WHITELIST or (len(t) >= 2 and TACTICAL_RE.search(t)):
                if t not in free:
                    free.append(t)
            elif t not in review:
                review.append(t)
    # 基础战术/违禁词兜底(保证折叠刀 等必然拦截)
    for t in ("防弹", "子弹", "折叠刀", "军刀", "匕首", "仿真枪", "弩", "甩棍", "指虎"):
        if t not in free:
            free.append(t)
    for t in ("毒品", "香烟", "枪支", "弹药", "炸药", "雷管", "剧毒", "易制毒", "管制刀具"):
        if t not in keywords:
            keywords.append(t)
    for t in review:
        report.append(f"[REVIEW] 战术列存疑词(未收录): {t}")
    # 去重可能错位的分类行(按 cat+sub+examples 组合)
    seen = set()
    for d in cats:
        key = (d["cat"], d["sub"], d["examples"])
        if key in seen:
            continue
        seen.add(key)
        ans = "【%s】%s。示例：%s。" % (d["cat"], d["limit"] or "禁售",
                                      d["examples"] or "—")
        _add(entries, seq, "rules", "%s类商品能上架吗？" % d["cat"],
             ans, "违禁规则", [d["cat"], "违禁", "上架"], "话术库/违禁自动刊登关键词",
             report)
    banned = {
        "keywords": keywords,
        "free_keywords": free,
        "note": "由 tools/build_corpus.py 从「违禁自动刊登关键词」表生成,含基础兜底词",
    }
    return entries, banned


def parse_pricing_fees(rows, report):
    """佣金、支付费:站点|佣金|交易手续费|基础建设费(空行间隔)。"""
    entries, seq = [], [0]
    CODE = {"菲律宾": "ph", "马来西亚": "my", "新加坡": "sg",
            "泰国": "th", "越南": "vn"}
    for row in rows[2:]:  # 前两行是表头 + 空行
        c = _cells(row)
        if not c or not c[0]:
            continue
        site, code = c[0], CODE.get(c[0], "")
        comm = c[1] if len(c) > 1 else ""
        fee = c[2] if len(c) > 2 else ""
        infra = c[3] if len(c) > 3 else ""
        if not comm and not fee and not infra:
            continue
        parts = []
        if comm:
            parts.append("佣金 %s" % comm)
        if fee:
            parts.append("交易手续费 %s" % fee)
        if infra:
            parts.append("基础建设费 %s" % infra)
        answer = "%s(%s)站：%s。" % (site, code.upper(), "，".join(parts))
        _add(entries, seq, "pricing",
             "%s佣金和交易手续费是多少？" % site, answer, "定价税费",
             [site, "佣金", "手续费"], "原表5/佣金、支付费", report,
             markets=[code] if code else None)
    return entries


def aggregate_category_fees(rows, report):
    """类目佣金率(1757 行 × 5 站点块)→ 只聚合每站 佣金 min/median/max 区间。"""
    entries, seq = [], [0]
    site_rates = defaultdict(list)
    for row in rows[1:]:
        c = _cells(row)
        for base in (0, 5, 10, 15, 20):
            if len(c) <= base + 3:
                continue
            site, rate = c[base + 2], c[base + 3]
            if site and rate:
                try:
                    site_rates[site].append(float(rate))
                except ValueError:
                    pass
    NAME = {"PH": "菲律宾", "MY": "马来西亚", "SG": "新加坡",
            "TH": "泰国", "VN": "越南"}
    for site, vals in sorted(site_rates.items()):
        vals.sort()
        mn, mx = vals[0], vals[-1]
        med = vals[len(vals) // 2]
        nm = NAME.get(site, site)
        ans = ("%s(%s)站：类目佣金率范围约 %.1f%%~%.1f%%,"
               "中位约 %.1f%%(仅区间,按类目差异大)。"
               % (nm, site, mn * 100, mx * 100, med * 100))
        _add(entries, seq, "pricing", "%s各品类佣金率大概多少？" % nm,
             ans, "定价税费", [nm, site, "佣金", "类目"], "原表5/类目佣金率(聚合)",
             report, markets=[site.lower()] if site.lower() in
             ("ph", "my", "sg", "th", "vn") else None)
    return entries


def aggregate_thai_tax(rows, report):
    """泰国税率(2204 行)→ 只聚合 VAT / 关税 / 综合费率区间,不出明细。"""
    duties, vats, comps = set(), set(), set()
    for row in rows[1:]:
        c = _cells(row)
        if len(c) < 7:
            continue
        for idx, col in ((4, duties), (5, vats), (6, comps)):
            try:
                col.add(float(c[idx]))
            except (ValueError, IndexError):
                pass
    if not comps:
        return []
    vat = min(vats) if vats else 0.0
    ans = ("泰国(TH)站：增值税(VAT)统一 %.0f%%;关税税率按类目 %.0f%%~%.0f%%;"
           "综合税费(关税+增值税复合)约 %.1f%%~%.1f%%。"
           % (vat * 100, min(duties) * 100 if duties else 0,
              max(duties) * 100 if duties else 0,
              min(comps) * 100, max(comps) * 100))
    return [{
        "id": "pricing-0001", "question": "泰国税费怎么算？",
        "answer": ans, "category": "定价税费",
        "tags": ["泰国", "税", "关税", "增值税", "TH"], "source": "原表5/泰国税率(聚合)",
        "lang": "zh", "markets": ["th"],
    }]


# ---------------------------------------------------------------- 主流程
def _dump_jsonl(path, entries):
    with open(path, "w", encoding="utf-8") as fh:
        for e in entries:
            fh.write(json.dumps(e, ensure_ascii=False) + "\n")
    return len(entries)


def build(src: Path, out_corpus: Path, out_config: Path, out_data: Path):
    out_corpus.mkdir(parents=True, exist_ok=True)
    out_data.mkdir(parents=True, exist_ok=True)
    report = []
    talk = src / TALKBOOK
    price = src / PRICING

    import openpyxl

    faq, rules_qa, pricing = [], [], []
    banned = {"keywords": [], "free_keywords": [],
              "note": "由 tools/build_corpus.py 生成"}

    def rows_of(book_path, name):
        """开 workbook 读某 sheet 全部行为元组列表后立即关闭(内存化)。"""
        wb = openpyxl.load_workbook(book_path, read_only=True, data_only=True)
        try:
            if name not in wb.sheetnames:
                return None
            return list(wb[name].iter_rows(values_only=True))
        finally:
            wb.close()

    # ---- 话术表 ----
    if talk.is_file():
        for sn, fn in (("快捷回复", parse_quick_reply),
                       ("常见问题助理", parse_faq_assistant),
                       ("客服的问题汇总", parse_customer_questions),
                       ("退货退款", parse_refund)):
            rows = rows_of(talk, sn)
            if rows is None:
                report.append(f"[SKIP] 话术表缺 sheet: {sn}")
                continue
            n = len(faq)
            faq.extend(fn(rows, report))
            report.append(f"[OK] {sn}: +{len(faq) - n} 条")
        rows = rows_of(talk, "违禁自动刊登关键词")
        if rows is not None:
            before = len(rules_qa)
            r, banned = parse_banned(rows, report)
            rules_qa.extend(r)
            report.append(f"[OK] 违禁自动刊登关键词: +{len(rules_qa) - before} 条规则, "
                          f"{len(banned['keywords'])} 违禁词 / {len(banned['free_keywords'])} 战术词")
    else:
        report.append(f"[MISS] 未找到话术表: {talk}")

    # ---- 定价税费表(只聚合) ----
    if price.is_file():
        wb = openpyxl.load_workbook(price, read_only=True, data_only=True)
        try:
            hit = [s for s in wb.sheetnames if s in FORBIDDEN_SHEETS]
            if hit:
                report.append(f"[ALERT] 定价表含禁读 sheet {hit},已跳过不解析")
        finally:
            wb.close()
        for sn, fn in (("佣金、支付费", parse_pricing_fees),
                       ("类目佣金率", aggregate_category_fees),
                       ("泰国税率", aggregate_thai_tax)):
            rows = rows_of(price, sn)
            if rows is None:
                report.append(f"[SKIP] 定价表缺 sheet: {sn}")
                continue
            n = len(pricing)
            pricing.extend(fn(rows, report))
            report.append(f"[OK] {sn}(聚合): +{len(pricing) - n} 条")
    else:
        report.append(f"[MISS] 未找到定价表: {price}")

    # ---- 写文件 ----
    n_faq = _dump_jsonl(out_corpus / "qa_faq.jsonl", faq)
    n_rules = _dump_jsonl(out_corpus / "qa_rules.jsonl", rules_qa)
    n_pricing = _dump_jsonl(out_corpus / "qa_pricing.jsonl", pricing)
    with open(out_config / "banned_keywords.json", "w", encoding="utf-8") as fh:
        json.dump(banned, fh, ensure_ascii=False, indent=2)
    dropped = [r for r in report if r.startswith("[DROP")]
    with open(out_data / "parse_report.md", "w", encoding="utf-8") as fh:
        fh.write("# 语料解析报告\n\n")
        fh.write("## 汇总\n\n")
        fh.write("- qa_faq.jsonl: %d 条\n- qa_rules.jsonl: %d 条\n"
                 "- qa_pricing.jsonl: %d 条\n- 违禁词: %d / 战术词: %d\n"
                 % (n_faq, n_rules, n_pricing,
                    len(banned["keywords"]), len(banned["free_keywords"])))
        fh.write("\n## 明细\n\n" + "\n".join(report) + "\n")
        fh.write("\n## 脱敏丢弃\n\n" + ("\n".join(dropped) or "(无)") + "\n")

    print(f"[build_corpus] faq={n_faq} rules={n_rules} pricing={n_pricing} "
          f"banned={len(banned['keywords'])} free={len(banned['free_keywords'])} "
          f"dropped={len(dropped)}", flush=True)
    print(f"[build_corpus] 报告: {out_data / 'parse_report.md'}", flush=True)


def main():
    ap = argparse.ArgumentParser(description="解析桌面 Excel 生成语料 + 违禁词表")
    ap.add_argument("--src", type=Path, default=DEFAULT_SRC,
                    help="Excel 所在目录(默认桌面新建文件夹)")
    args = ap.parse_args()
    build(args.src, CORPUS_DIR, CONFIG_DIR, DATA_DIR)


if __name__ == "__main__":
    main()
